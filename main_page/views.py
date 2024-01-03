import os
import pandas as pd
from datetime import datetime
from urllib.parse import urlparse
from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.urls import resolve
from django.template import loader
from django.conf import settings
from .utils import (
    load_dataframe_from_file, save_dataframe,
    dataframe_to_html, record_action, 
    get_actions_for_session, save_as_template, action_to_dict,
    handle_remove_empty_rows, handle_remove_empty_cols,
    handle_delete_first_x_rows, handle_delete_last_x_rows,
    handle_replace_header_with_first_row, add_column,
    delete_columns, fill_column, split_column,
    merge_columns, rename_column, trim_and_replace_multiple_whitespaces, delete_data, replace_symbol, change_case
)
from .forms import UploadFileForm, ParagraphErrorList
from .models import Action, UploadedFile, Template
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.views.decorators.http import require_POST
import re
import json
import openpyxl
from openpyxl import load_workbook


def main_page(request):
    context = {
        'previous_page_url': None,
        'next_page_url': 'summary',
        'active_page': 'home',
        'multiple_sheets': False,
        'sheet_names': [],
        'temp_file_path': request.session.get('temp_file_path', None),
        'merged_cells': False,
    }
    if request.method == 'POST':
        if 'file_upload' in request.POST:
            form = UploadFileForm(request.POST, request.FILES, error_class=ParagraphErrorList)
            if form.is_valid():
                uploaded_file = form.save()  # saves the original file
                file_path = uploaded_file.file.path
                request.session['file_path'] = file_path
                request.session['temp_file_path'] = file_path
                print(f"file path after file upload: {file_path}")

                # checks if the file is excel file and has multiple sheets
                if file_path.endswith(('.xlsx', '.xls')):
                    workbook = openpyxl.load_workbook(file_path)
                    sheet = workbook.active
                    if sheet.merged_cells.ranges:
                        context['merged_cells'] = True
                        context['form'] = form
                        return render(request, '1_index.html', context)

                    xls = pd.ExcelFile(file_path)
                    sheet_names = xls.sheet_names
                    if len(sheet_names) > 1:
                        context['multiple_sheets'] = True
                        context['sheet_names'] = sheet_names
                        return render(request, '1_index.html', context)
                    else:
                        df_orig = pd.read_excel(file_path, sheet_name=sheet_names[0])

                else:
                    df_orig = load_dataframe_from_file(file_path)
                
                # creates a modified copy
                file_dir, file_name = os.path.split(file_path)
                file_root, _ = os.path.splitext(file_name)
                temp_file_path = os.path.join(file_dir, f"TEMP_{file_root}.csv")
                
                save_dataframe(df_orig, temp_file_path, file_format='csv')
                df_orig = load_dataframe_from_file(temp_file_path)
                html_table = dataframe_to_html(df_orig,classes='table table-striped preserve-whitespace')
                request.session['html_table'] = html_table
                request.session['temp_file_path'] = temp_file_path
                print(f"file path after saving dataframe: {file_path}")
                request.session.save()
                return redirect('summary')
                
            else:
                context['form'] = form  # adds the invalid form to the context so errors can be displayed
                return render(request, '1_index.html', context)
        elif 'sheet_selection' in request.POST:
            file_path = request.session.get('temp_file_path')
            if file_path:
                selected_sheet = request.POST.get('sheet')

                # checks if the file is Excel and appends or selects sheets accordingly
                if file_path.endswith(('.xlsx', '.xls')):
                    xls = pd.ExcelFile(file_path)

                    if selected_sheet == '__append_all__':
                        # appends all sheets
                        df_list = [pd.read_excel(xls, sheet_name=sheet) for sheet in xls.sheet_names]
                        df_orig = pd.concat(df_list, ignore_index=True)
                    else:
                        # loads the selected sheet
                        df_orig = pd.read_excel(xls, sheet_name=selected_sheet)

                    temp_file_dir, file_name = os.path.split(file_path)
                    file_root, _ = os.path.splitext(file_name)
                    temp_file_path = os.path.join(temp_file_dir, f"TEMP_{file_root}.csv")
                    df_orig.to_csv(temp_file_path, index=False)

                    # updates the session with the new file path
                    request.session['temp_file_path'] = temp_file_path
                    request.session['html_table'] = df_orig.to_html(classes='table table-striped preserve-whitespace')

                else:
                    pass

                file_dir, file_name = os.path.split(file_path)
                file_root, _ = os.path.splitext(file_name)
                temp_file_path = os.path.join(file_dir, f"TEMP_{file_root}.csv")
                
                save_dataframe(df_orig, temp_file_path, file_format='csv')
                df_orig = load_dataframe_from_file(temp_file_path)
                html_table = dataframe_to_html(df_orig,classes='table table-striped preserve-whitespace')
                request.session['html_table'] = html_table
                request.session['temp_file_path'] = temp_file_path
                print(f"file path after saving dataframe: {file_path}")
                request.session.save()
                return redirect('summary')
                
        elif 'merged_cell_action' in request.POST:
            file_path = request.session.get('temp_file_path')
            merged_cell_action = request.POST.get('merged_cell_action')

            if file_path and file_path.endswith(('.xlsx', '.xls')):
                workbook = load_workbook(file_path, data_only=True)
                sheet = workbook.active

                if merged_cell_action == 'first_cell':
                    # default behaviour
                    pass
                elif merged_cell_action == 'all_cells':
                    unmerge_and_fill_all_cells(sheet)
                elif merged_cell_action == 'delete_rows':
                    delete_rows_with_merged_cells(sheet)

                df_orig = pd.DataFrame(sheet.values)
                df_orig.columns = df_orig.iloc[0]
                df_orig = df_orig[1:]
                file_dir, file_name = os.path.split(file_path)
                file_root, _ = os.path.splitext(file_name)
                temp_file_path = os.path.join(file_dir, f"TEMP_{file_root}.csv")
                
                save_dataframe(df_orig, temp_file_path, file_format='csv')
                df_orig = load_dataframe_from_file(temp_file_path)
                html_table = dataframe_to_html(df_orig,classes='table table-striped preserve-whitespace')
                request.session['html_table'] = html_table
                request.session['temp_file_path'] = temp_file_path
                print(f"file path after saving dataframe: {file_path}")
                request.session.save()
                return redirect('summary')

    else:
        form = UploadFileForm()
        context['form'] = form
        return render(request, '1_index.html', context)

    


def summary(request):
    print("Entered summary view")
    temp_file_path = request.session.get('temp_file_path')
    file_path = request.session.get('file_path')
    df_v1 = load_dataframe_from_file(temp_file_path)
    print(f"file path in summary view: {file_path}")
    print(f"temp file path in summary view: {temp_file_path}")
        # checks if a file has been uploaded
    if not temp_file_path or not file_path:
        messages.error(request, "No file uploaded. Please upload a file to proceed.")
        return redirect('main_page')
    # extracts the relative path using MEDIA_ROOT
    relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT).replace('\\', '/')
    print(relative_path)
        # checks if the user is authenticated before querying the database
    if request.user.is_authenticated:
        user_templates = Template.objects.filter(user=request.user)
    else:
        user_templates = Template.objects.none()

    # retrieves data from session
    header_mismatch = request.session.get('header_mismatch', False)
    mismatched_headers = request.session.get('mismatched_headers', ([], []))
    mismatched_headers_marked = ([], [])

    template_application_success = False

    try:
        uploaded_file_instance = UploadedFile.objects.get(file=relative_path)
    except UploadedFile.DoesNotExist:
        # handles the case where the UploadedFile instance doesn't exist
        messages.error(request, "The file you're working on could not be found.")
        return redirect('main_page')  # redirects to a safe page
    # identifies emty columns
    empty_cols = df_v1.columns[df_v1.isna().all()].tolist()

    if request.method == 'POST':
        print("POST request received in summary view")
        # deletes all empty rows
        if 'remove_empty_rows' in request.POST:
            print("remove_empty_rows action detected")
            try:
                result = handle_remove_empty_rows(df_v1)
                df_v1 = result['dataframe']
                rows_removed = result['rows_removed']
                messages.success(request, f'{rows_removed} empty rows removed.')
            except ValueError as e:
                messages.error(request, str(e))
            record_action(
                uploaded_file=uploaded_file_instance,        
                action_type='remove_empty_rows',
                parameters={},
                user=request.user,
                session_id=request.session.session_key,
                )
           
        # deletes selected columns
        if 'remove_empty_cols' in request.POST:
            try:
                cols_to_delete = request.POST.getlist('remove_empty_cols')
                result = handle_remove_empty_cols(df_v1, cols_to_delete)
                df_v1 = result['dataframe']
                cols_removed = result['cols_removed']

                messages.success(request, f'{cols_removed} empty columns removed.')
            except ValueError as e:
                messages.error(request, str(e))
            record_action(        
                    uploaded_file=uploaded_file_instance,
                    action_type='remove_empty_cols',
                    parameters={'cols_to_delete': cols_to_delete},
                    user=request.user,
                    session_id=request.session.session_key,
                    )

        if 'num_rows_to_delete_start' in request.POST:
            try:
                num_rows_to_delete_start = request.POST.get('num_rows_to_delete_start', '0').strip()
                if num_rows_to_delete_start.isdigit():
                    num_rows_to_delete_start = int(num_rows_to_delete_start)
                else:
                    num_rows_to_delete_start = 0
                if num_rows_to_delete_start > 0:
                    result = handle_delete_first_x_rows(df_v1, num_rows_to_delete_start)
                    df_v1 = result['dataframe']
                    rows_deleted = result['rows_deleted']
                    messages.success(request, f'First {rows_deleted} rows deleted successfully.')

                    record_action(        
                            uploaded_file=uploaded_file_instance,
                            action_type='delete_first_X_rows',
                            parameters={'num_rows_to_delete_start': num_rows_to_delete_start},
                            user=request.user,
                            session_id=request.session.session_key,
                            )
            except ValueError as e:
                messages.error(request, str(e))
        
        # if replace_header is True, sets the df columns to the first row's values
        if 'replace_header' in request.POST:
            try:
                df_v1 = handle_replace_header_with_first_row(df_v1)
                messages.success(request, 'Header replaced with the first row successfully.')

                record_action(        
                        uploaded_file=uploaded_file_instance,
                        action_type='replace_header',
                        parameters={},
                        user=request.user,
                        session_id=request.session.session_key,
                        )
            except ValueError as e:
                messages.error(request, str(e))

        if 'num_rows_to_delete_end' in request.POST:
            try:
                num_rows_to_delete_end = request.POST.get('num_rows_to_delete_end', '0').strip()
                if num_rows_to_delete_end.isdigit():
                    num_rows_to_delete_end = int(num_rows_to_delete_end)
                else:
                    num_rows_to_delete_end = 0
                if num_rows_to_delete_end > 0:
                    result = handle_delete_last_x_rows(df_v1, num_rows_to_delete_end)
                    df_v1 = result['dataframe']
                    rows_deleted = result['rows_deleted']
                    messages.success(request, f'Last {rows_deleted} rows deleted successfully.')

                    record_action(        
                            uploaded_file=uploaded_file_instance,
                            action_type='delete_last_X_rows',
                            parameters={'num_rows_to_delete_end': num_rows_to_delete_end},
                            user=request.user,
                            session_id=request.session.session_key,
                            )
            except ValueError as e:
                messages.error(request, str(e))
        
        elif 'apply_template' in request.POST:
            template_id = request.POST.get('template_id')
            try:
                template = Template.objects.get(id=template_id, user=request.user)
                df_original = load_dataframe_from_file(temp_file_path)

                current_headers = df_original.columns.tolist()
                print(f"current headers: {current_headers}, original headers: {template.original_headers}")

                if set(current_headers) != set(template.original_headers):
                    header_mismatch = True
                    print("header mismatch")
                    print(f"{header_mismatch}")
                    
                    mismatched_headers = (current_headers, template.original_headers)
                    print(f"{mismatched_headers}")

                    request.session['header_mismatch'] = header_mismatch
                    request.session['mismatched_headers'] = mismatched_headers
                    
                else:
                    print("headers match")
                    mismatched_headers_marked = ([], [])
                    print("Applying actions from template:", template.actions) 
                    for action in template.actions:
                        action_type = action['action_type']
                        parameters = action['parameters']
                        df_original, error_message = apply_action(df_original, action_type, parameters)
                        if error_message:
                            # if there's an error, displays it and reverts to the original df
                            messages.error(request, f"Error applying template: {error_message}")
                            save_dataframe(df_original, temp_file_path)  # saves the original df
                            return redirect('summary')
                    
                    df_v1 = df_original
                    save_dataframe(df_original, temp_file_path)
                    template_application_success = True
                    messages.success(request, "Template applied successfully.")

            except Template.DoesNotExist:
                messages.error(request, "Template not found or access denied.")

        if 'adjust_columns' in request.POST:
            new_columns = request.POST.getlist('new_column[]')
            columns_to_delete = request.POST.getlist('delete_columns[]')
            
            # adding new columns
            for col in new_columns:
                if col:  # Add column if name is provided
                    df_v1[col] = None

            # removing selected columns
            df_v1.drop(columns=columns_to_delete, errors='ignore', inplace=True)
            
            save_dataframe(df_v1, temp_file_path)
            messages.success(request, "Columns adjusted successfully.")

        # map headers logic
        if 'map_headers' in request.POST:
            for header in df_v1.columns:
                print(f'old header-{header}')
                new_header = request.POST.get(f'header-{header}')
                print(f'new header-{new_header}')
                if new_header:
                    df_v1.rename(columns={header: new_header}, inplace=True)

            save_dataframe(df_v1, temp_file_path)
            messages.success(request, "Headers mapped successfully.")
        print("About to save DataFrame")
        temp_file_path = save_dataframe(df_v1, temp_file_path)
        print(f"DataFrame saved to: {temp_file_path}")
        # updates the session with the new file path
        request.session['temp_file_path'] = temp_file_path
        # redirects to avoid resubmit on refresh
        return redirect('summary')
    
        # clears the session values if they are not needed anymore
    if header_mismatch:
        current_mismatched = set(mismatched_headers[0])
        template_mismatched = set(mismatched_headers[1])
        diff_headers = current_mismatched.symmetric_difference(template_mismatched)

        current_headers_marked = [(header, header in diff_headers) for header in mismatched_headers[0]]
        template_headers_marked = [(header, header in diff_headers) for header in mismatched_headers[1]]

        mismatched_headers_marked = (current_headers_marked, template_headers_marked)
        # Clear the session values
        del request.session['header_mismatch']
        del request.session['mismatched_headers']

    context = {
        'num_rows': df_v1.shape[0],
        'num_cols': df_v1.shape[1],
        'col_names': df_v1.columns.tolist(),
        'num_empty_rows': df_v1[df_v1.isna().all(axis=1)].shape[0],
        'num_empty_cols': df_v1.columns[df_v1.isna().all(axis=0)].size,
        'empty_cols': empty_cols,
        'header_mismatch': header_mismatch,
        'mismatched_headers': mismatched_headers,
        'mismatched_headers_marked': mismatched_headers_marked,
        'template_application_success': template_application_success,
        'table': dataframe_to_html(df_v1, classes='table table-striped preserve-whitespace'),
        'original_file_name': os.path.basename(file_path),
        'templates': user_templates,
        'previous_page_url': 'main_page',
        'active_page': 'summary',
        'next_page_url': 'edit_columns',
    }
    
    return render(request, '2_file_summary.html', context)

def edit_columns(request):
    temp_file_path = request.session.get('temp_file_path')
    file_path = request.session.get('file_path')
        # checks if a file has been uploaded
    if not temp_file_path or not file_path:
        messages.error(request, "No file uploaded. Please upload a file to proceed.")
        return redirect('main_page')
    # extracts the relative path using MEDIA_ROOT
    relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT).replace('\\', '/')

    try:
        uploaded_file_instance = UploadedFile.objects.get(file=relative_path)
    except UploadedFile.DoesNotExist:
        # handles the case where the UploadedFile instance doesn't exist
        messages.error(request, "The file you're working on could not be found.")
        return redirect('main_page')  # redirects to a safe page
    df_v2 = load_dataframe_from_file(temp_file_path)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add_column':
            new_column_name = request.POST.get('new_column_name')
            print(f"New column name: {new_column_name}")  
            if new_column_name:
                df_v2 = add_column(df_v2, new_column_name)
            messages.success(request, f'Column "{new_column_name}" added successfully.')
            record_action(
                uploaded_file=uploaded_file_instance,        
                action_type='add_column',
                parameters={'new_column_name': new_column_name},
                user=request.user,
                session_id=request.session.session_key,
            )


        elif action == 'delete_columns':
            columns_to_delete = request.POST.getlist('columns_to_delete')
            df_v2 = delete_columns(df_v2, columns_to_delete)
            messages.success(request, f'{len(columns_to_delete)} columns deleted successfully.')
            record_action(
                uploaded_file=uploaded_file_instance,        
                action_type='delete_columns',
                parameters={'columns_to_delete': columns_to_delete},
                user=request.user,
                session_id=request.session.session_key,
            )


        elif action == 'fill_column':
            column_to_fill = request.POST.get('column_to_fill')
            fill_value = request.POST.get('fill_value')
            fill_option = request.POST.get('fill_option')

            if column_to_fill and fill_value is not None:
                df_v2 = fill_column(df_v2, column_to_fill, fill_value, fill_option)
            messages.success(request, f'Column "{column_to_fill}" filled successfully with "{fill_value}".')

            record_action(
                uploaded_file=uploaded_file_instance,        
                action_type='fill_column',
                parameters={
                    'column_to_fill': column_to_fill,
                    'fill_value': fill_value,
                    'fill_option': fill_option
                },
                user=request.user,
                session_id=request.session.session_key,
            )


        elif action == 'split_column':
            column_to_split = request.POST.get('column_to_split')
            split_value = request.POST.get('split_value')
            delete_original = 'delete_original' in request.POST
            ignore_repeated = 'ignore_repeated' in request.POST

            if column_to_split in df_v2.columns and split_value:
                # performs the split operation
                df_v2 = split_column(df_v2, column_to_split, split_value, delete_original, ignore_repeated)
            messages.success(request, f'Column "{column_to_split}" split successfully.')
            record_action(
                uploaded_file=uploaded_file_instance,        
                action_type='split_column',
                parameters={
                    'column_to_split': column_to_split,
                    'split_value': split_value,
                    'delete_original': delete_original,
                    'ignore_repeated': ignore_repeated
                },
                user=request.user,
                session_id=request.session.session_key,
            )


        elif action == 'merge_columns':
            columns_to_merge = request.POST.getlist('columns_to_merge')
            merge_separator = request.POST.get('merge_separator', '')
            new_column_name = request.POST.get('new_merge_column_name')
            delete_original = 'delete_original_after_merge' in request.POST

            merge_result = merge_columns(df_v2, columns_to_merge, merge_separator, new_column_name, delete_original)
            df_v2 = merge_result['dataframe']  # updates df
            used_column_name = merge_result['new_column_name']  # extracts the new column name

            merge_message = f'Columns {", ".join(columns_to_merge)} merged into "{used_column_name}" successfully.'
            if delete_original:
                merge_message += ' Original columns have been deleted.'
            messages.success(request, merge_message)
            record_action(
                uploaded_file=uploaded_file_instance,        
                action_type='merge_columns',
                parameters={
                    'columns_to_merge': columns_to_merge,
                    'merge_separator': merge_separator,
                    'new_column_name': used_column_name,
                    'delete_original': delete_original
                },
                user=request.user,
                session_id=request.session.session_key,
            )


        elif action == 'rename_column':
            column_to_rename = request.POST.get('column_to_rename')
            new_column_name = request.POST.get('new_renamed_column_name')

            if column_to_rename in df_v2.columns and new_column_name:
                df_v2 = rename_column(df_v2, column_to_rename, new_column_name)
            messages.success(request, f'Column "{column_to_rename}" renamed to "{new_column_name}" successfully.')
            record_action(
                uploaded_file=uploaded_file_instance,        
                action_type='rename_column',
                parameters={
                    'column_to_rename': column_to_rename,
                    'new_column_name': new_column_name
                },
                user=request.user,
                session_id=request.session.session_key,
            )

            
        save_dataframe(df_v2, temp_file_path)
        return redirect('edit_columns')

    context = {
        'previous_page_url': 'summary',
        'active_page': 'edit_columns',
        'next_page_url': 'edit_data',
        'table': dataframe_to_html(df_v2, classes='table table-striped preserve-whitespace'),
        'original_file_name': os.path.basename(file_path),
        'df_v2': df_v2,
    }
    return render(request, '3_edit_columns.html', context)

def edit_data(request):

    temp_file_path = request.session.get('temp_file_path')
    file_path = request.session.get('file_path')
    # checks if a file has been uploaded
    if not temp_file_path or not file_path:
        messages.error(request, "No file uploaded. Please upload a file to proceed.")
        return redirect('main_page')
    # extracts the relative path using MEDIA_ROOT
    relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT).replace('\\', '/')

    try:
        uploaded_file_instance = UploadedFile.objects.get(file=relative_path)
    except UploadedFile.DoesNotExist:
        # handles the case where the UploadedFile instance doesn't exist
        messages.error(request, "The file you're working on could not be found.")
        return redirect('main_page')  # redirects to a safe page
    df_v3 = load_dataframe_from_file(temp_file_path)
    if request.method == 'POST':
        print("POST request received")
        action = request.POST.get('action')
        print(f"Action received: {action}")  # prints the action value

        if action == 'delete_data':
            columns_to_modify = request.POST.getlist('columns_to_modify')
            delimiter = request.POST.get('delimiter')
            delete_option = request.POST.get('delete_option')
            include_delimiter = 'include_delimiter' in request.POST
            apply_to_all = '__all__' in columns_to_modify
            case_sensitive = 'case_sensitive' in request.POST

            if apply_to_all:
                columns_to_modify = df_v3.columns.tolist()  # lists all columns if '--ALL COLUMNS--' is selected

            try:
                df_v3 = delete_data(df_v3, columns_to_modify, delimiter, delete_option, include_delimiter, case_sensitive)
                messages.success(request, 'Data deleted successfully based on your criteria.')
            except ValueError as e:
                messages.error(request, str(e))
            df_backup = df_v3[columns_to_modify].copy()  # backups the original data
            record_action(
                uploaded_file=uploaded_file_instance,        
                action_type='delete_data',
                parameters={
                    'columns_to_modify': columns_to_modify,
                    'delimiter': delimiter,
                    'delete_option': delete_option,
                    'include_delimiter': include_delimiter,
                    'case_sensitive': case_sensitive
                },
                user=request.user,
                session_id=request.session.session_key,
                df_backup=df_backup,
            )



        elif action == 'replace_symbol':
            columns_to_replace = request.POST.getlist('columns_to_replace')
            old_symbol = request.POST.get('old_symbol')
            new_symbol = request.POST.get('new_symbol')
            # defaults to False if the checkbox is not checked
            case_sensitive = 'case_sensitive' in request.POST
            apply_to_all = '__all__' in columns_to_replace

            if apply_to_all:
                columns_to_replace = df_v3.columns.tolist()  # lists all columns if '--ALL COLUMNS--' is selected

            try:
                df_v3 = replace_symbol(df_v3, columns_to_replace, old_symbol, new_symbol, case_sensitive)
                messages.success(request, 'Symbols replaced successfully.')
            except ValueError as e:
                messages.error(request, str(e))

           
            record_action(
                uploaded_file=uploaded_file_instance,        
                action_type='replace_symbol',
                parameters={
                    'columns_to_replace': columns_to_replace if not apply_to_all else 'All Columns',
                    'old_symbol': old_symbol,
                    'new_symbol': new_symbol,
                    'case_sensitive': case_sensitive
                },
                user=request.user,
                session_id=request.session.session_key,
            )

            

        elif action == 'validate_data':
            columns_to_validate = request.POST.getlist('columns_to_validate')
            validation_type = request.POST.get('validation_type')
            regex_pattern = request.POST.get('regex_pattern')
            ignore_whitespace = 'ignore_whitespace' in request.POST
            ignore_empty = 'ignore_empty' in request.POST
            apply_to_all = '__all__' in columns_to_validate

            print("Validation Type:", validation_type)
            print("Ignore Whitespace:", ignore_whitespace)
            print("Ignore Empty:", ignore_empty)
            print("Apply to All Columns:", apply_to_all)

            if apply_to_all:
                columns_to_validate = df_v3.columns.tolist()

            invalid_rows = {}
            total_invalid_rows = 0

            for column in columns_to_validate:
                if column in df_v3.columns:
                    column_series = df_v3[column].astype(str)
                    invalid_rows[column] = []

                    for index, value in column_series.items():
                        print(f"Validating column '{column}', row {index}, value: '{value}'")

                        # checks for empty cells based on ignore_empty flag
                        if pd.isna(df_v3.at[index, column]):
                            if not ignore_empty:
                                invalid_rows[column].append(index)
                            continue  # skips further validation for this cell

                        if ignore_whitespace:
                            value = value.replace(' ', '')

                        is_invalid = False
                        if validation_type == 'letters' and not value.isalpha():
                            is_invalid = True
                        elif validation_type == 'numbers' and not value.isdigit():
                            is_invalid = True
                        elif validation_type == 'no_specials' and not value.isalnum():
                            is_invalid = True
                        elif validation_type == 'regex' and not re.match(regex_pattern, value):
                            is_invalid = True

                        if is_invalid:
                            invalid_rows[column].append(index)

                    total_invalid_rows += len(invalid_rows[column])

            print(f"Total invalid rows: {total_invalid_rows}")
            print("Invalid rows details:", invalid_rows)

            # Constructing the message
            if total_invalid_rows > 0:
                invalid_message = f"Validation failed for {total_invalid_rows} rows. \n"
                for column, rows in invalid_rows.items():
                    if rows:
                        invalid_message += f"In column '{column}': Rows {', '.join(map(str, rows))}\n"
                messages.error(request, invalid_message)
            else:
                messages.success(request, "All rows validated successfully.")

            print(f"Invalid rows: {invalid_rows}")

        elif action == 'check_duplicates':
            columns_to_check = request.POST.getlist('columns_to_check_duplicates')
            if columns_to_check:
                if '__all__' in columns_to_check:
                    columns_to_check = df_v3.columns.tolist()

                duplicates = df_v3[df_v3.duplicated(subset=columns_to_check, keep=False)]
                duplicates.sort_values(by=columns_to_check, inplace=True)


                # constructs the duplicate message
                if not duplicates.empty:
                    duplicate_rows = duplicates.index.tolist()
                    message = f"Found {len(duplicates)} duplicate rows: {', '.join(map(str, duplicate_rows))}."
                    messages.error(request, message)
                else:
                    messages.success(request, "No duplicates found.")


        elif action == 'trim_and_replace_whitespaces':
            columns_to_modify = request.POST.getlist('columns_to_trim')
            apply_to_all = '__all__' in columns_to_modify
            try:
                df_v3 = trim_and_replace_multiple_whitespaces(df_v3, columns_to_modify, replace_all=apply_to_all)
                messages.success(request, 'Unnecessary whitespaces replaced successfully.')
            except ValueError as e:
                messages.error(request, str(e))
            record_action(
                uploaded_file=uploaded_file_instance,
                action_type='trim_and_replace_whitespaces',
                parameters={
                    'columns_to_modify': columns_to_modify if not apply_to_all else 'All Columns'
                },
                user=request.user,
                session_id=request.session.session_key
            )


        elif action == 'change_case':
            columns_to_change_case = request.POST.getlist('columns_to_change_case')
            case_type = request.POST.get('case_type')
            apply_to_all = '__all__' in columns_to_change_case

            if apply_to_all:
                columns_to_change_case = df_v3.columns.tolist()

            try:
                df_v3 = change_case(df_v3, columns_to_change_case, case_type)
                messages.success(request, 'Case changed successfully.')
            except ValueError as e:
                messages.error(request, str(e))

            record_action(
                uploaded_file=uploaded_file_instance,
                action_type='change_case',
                parameters={
                    'columns_to_change_case': columns_to_change_case if not apply_to_all else 'All Columns',
                    'case_type': case_type
                },
                user=request.user,
                session_id=request.session.session_key
            )



       

        save_dataframe(df_v3, temp_file_path)
        return redirect('edit_data')
    
    context = {
        'previous_page_url': 'edit_columns',
        'active_page': 'edit_data',
        'next_page_url': 'download',
        'table': dataframe_to_html(df_v3, classes='table table-striped preserve-whitespace'),
        'original_file_name': os.path.basename(file_path),
        'df_v3': df_v3,
        'duplicates_json': request.session.get('duplicates_json', '[]'),  # pass the duplicates as JSON
    }
    
    return render(request, '4_edit_data.html', context)




def download(request):
    temp_file_path = request.session.get('temp_file_path')  
    file_path = request.session.get('file_path')

    # checks if a file has been uploaded
    if not temp_file_path or not file_path:
        messages.error(request, "No file uploaded. Please upload a file to proceed.")
        return redirect('main_page')


    context = {
        'original_file_name': os.path.basename(file_path),
        'previous_page_url': 'edit_columns',
        'active_page': 'download',
        'next_page_url': None  # to disable pagination
    }
    file_format = request.GET.get('format')
    print(f"File format: {file_format}")

    
    if file_format:
        
        original_file_name = os.path.basename(file_path)
        original_file_dir = os.path.dirname(file_path)  # get the directory of the original file
        
        print(f"Original file name: {original_file_name}")
        df = load_dataframe_from_file(temp_file_path)
        print(df.head())
        file_name_no_ext, _ = os.path.splitext(original_file_name)

        # generate the new filename
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        new_file_name = f"{file_name_no_ext}_EDITED_{timestamp}"
        print(f"New file name: {new_file_name}")
        save_path = os.path.join(original_file_dir, f"{new_file_name}.{file_format}")  # save in the original file's directory
        save_dataframe(df, save_path, file_format)

        with open(save_path, 'rb') as f:
            if file_format == 'csv':
                response = HttpResponse(f, content_type='text/csv')
            elif file_format == 'xlsx':
                response = HttpResponse(f, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            elif file_format == 'json':
                response = HttpResponse(f, content_type='application/json')
            elif file_format == 'xml':
                response = HttpResponse(f, content_type='application/xml')
            elif file_format == 'tsv':
                response = HttpResponse(f, content_type='text/tab-separated-values')
        
            else:
                # handle unexpected format
                return HttpResponse("Unexpected format", status=400)
                
        response['Content-Disposition'] = f'attachment; filename="{new_file_name}.{file_format}"'
        return response
    else:
        # if no format is specified, render the HTML page
        return render(request, '5_download.html', context)




@login_required(login_url="/login/")
def user_profile(request):
    user_templates = Template.objects.filter(user=request.user)  # retrieves templates for the current user

    context = {
        'user': request.user,
        'templates': user_templates,
        'date_joined': request.user.date_joined, 
        'last_login': request.user.last_login,    
        'active_page': 'user_profile', # for highlighting the active page in the navbar
    }
    return render(request, 'user_profile.html', context)



def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            messages.success(request, f'Account created for {username}! You can now log in.')
            return redirect('login')
    else:
        form = UserCreationForm()
    return render(request, 'register.html', {'form': form})

@login_required(login_url="/login/")
def adjust_columns(request):
    temp_file_path = request.session.get('temp_file_path')
    if request.method == 'POST':
        df = load_dataframe_from_file(temp_file_path)

        # Adding new columns
        new_columns = request.POST.getlist('new_column[]')
        for col in new_columns:
            if col:  # adds column if name is provided
                df[col] = None

        # Removing selected columns
        columns_to_delete = request.POST.getlist('delete_columns[]')
        df.drop(columns=columns_to_delete, errors='ignore', inplace=True)

        save_dataframe(df, temp_file_path)

    return redirect('summary')


@login_required(login_url="/login/")
def map_headers(request):
    temp_file_path = request.session.get('temp_file_path')
    if request.method == 'POST':
        df = load_dataframe_from_file(temp_file_path)

        # maps headers based on user input
        for header in df.columns:
            new_header = request.POST.get(f'header-{header}')
            if new_header:
                df.rename(columns={header: new_header}, inplace=True)

        save_dataframe(df, temp_file_path)

    return redirect('summary')

@csrf_exempt
def delete_file(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        if data.get('delete'):
            file_path = request.session.get('file_path')
            if file_path:
                try:
                    os.remove(file_path)
                    del request.session['file_path']  # removes file path from session
                    return JsonResponse({'status': 'success'}, status=200)
                except Exception as e:
                    return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
            else:
                return JsonResponse({'status': 'error', 'message': 'File path not found.'}, status=404)
    return JsonResponse({'status': 'error', 'message': 'Invalid request.'}, status=400)

@login_required
@require_POST
@csrf_protect
def delete_template(request, template_id):
    # gets the template object, ensuring it belongs to the current user
    template = get_object_or_404(Template, id=template_id, user=request.user)

    try:
        template.delete()
        messages.success(request, 'Template deleted successfully.')
        # redirects after successful deletion
        return redirect('user_profile')
    except Exception as e:
        messages.error(request, f'Error deleting template: {str(e)}')
        # redirects after failure
        return redirect('user_profile')

def unmerge_and_fill_all_cells(sheet):
    # vreates a list of merged cell ranges before modifying them
    merged_ranges = list(sheet.merged_cells.ranges)

    for merged_cell_range in merged_ranges:
        top_left_cell = sheet.cell(row=merged_cell_range.min_row, column=merged_cell_range.min_col)
        first_cell_value = top_left_cell.value
        sheet.unmerge_cells(str(merged_cell_range))
        for row in sheet.iter_rows(min_row=merged_cell_range.min_row, max_row=merged_cell_range.max_row,
                                   min_col=merged_cell_range.min_col, max_col=merged_cell_range.max_col):
            for cell in row:
                cell.value = first_cell_value

def delete_rows_with_merged_cells(sheet):
    rows_to_delete = set()
    for merged_cell_range in sheet.merged_cells.ranges:
        for row in range(merged_cell_range.min_row, merged_cell_range.max_row + 1):
            rows_to_delete.add(row)
    for row in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row)

def undo_last_action(request):
    print("Undo view accessed")
    # determines the current view based on the request path, default to summary
    current_view = request.POST.get('current_view', 'summary')
    original_file_path = request.session.get('file_path')

    if original_file_path is None:
        print("Original file path is None!")
        return JsonResponse({'status': 'error', 'error': 'Original file path not found'}, status=500)

    print(f"Original file path: {original_file_path}")

    # extracts the relative path using MEDIA_ROOT
    relative_path = os.path.relpath(original_file_path, settings.MEDIA_ROOT).replace('\\', '/')
    print(f"Relative path: {relative_path}")

    try:
        current_file = UploadedFile.objects.get(file=relative_path)
        print("UploadedFile instance found:", current_file)
    except UploadedFile.DoesNotExist:
        print("UploadedFile instance not found.")
        messages.error(request, "The file you're working on could not be found.")
        return redirect('main_page')

    # retrieves the df from the file
    df = load_dataframe_from_file(original_file_path)
    print("DataFrame loaded from file.")

    # retrieves actions that are not undone yet
    actions = get_actions_for_session(request.session.session_key, current_file, exclude_last_action=False)
    print(f"Actions retrieved: {actions}")

    if actions.exists():
        # marks the last action as undone
        last_action = actions.latest('timestamp')

        action_failed = False


        # apply the actions
        for action in actions:
            try:
                # skips applying the last action
                if action.id == last_action.id:
                    continue
                df, current_view = apply_action(df, action.action_type, action.parameters, is_undo=True)
                print(f"Applied action: {action.action_type}, Current view: {current_view}")
                if current_view is None:
                    break
            except Exception as e:
                messages.error(request, f"An error occurred while undoing the last action: {str(e)}")
                action_failed = True
                break
        
        if not action_failed and last_action:
        # marks the last action as undone only if no errors occurred in applying actions
            last_action.undone = True
            last_action.save()
            print("Last action marked as undone.")
        # retrieves action name and parameters
        action_name = last_action.action_type.replace('_', ' ').capitalize()
        action_params = last_action.parameters if last_action.parameters else {}

        # formats parameters for display
        formatted_params = ', '.join([f'{key}: {value}' for key, value in action_params.items()])
        message = f"Successfully undone action: '{action_name}' with parameters ({formatted_params})."
        messages.success(request, message)
    else:
        messages.info(request, "No more actions to undo.")
        print("No actions to undo.")
        return redirect(current_view)

    # checks if current_view is not None before redirecting
    if current_view is not None:
        # saves the modified df back to the temporary file path
        temp_file_path = request.session.get('temp_file_path')
        save_dataframe(df, temp_file_path)
        print("DataFrame saved, redirecting to current view:", current_view)
        return redirect(current_view)
    else:
        print("An error occurred while undoing the last action.")
        messages.error(request, "An error occurred while undoing the last action.")
        return redirect(current_view)


def apply_action(df, action_type, parameters, is_undo=False):
    try:
        print(f"Applying {action_type} with {parameters}")
        # prints initial state of df
        print("DataFrame before action:", df.head())
        current_view = None  # variable to hold the current view name

        # defines view names for each action type
        summary_view_actions = ['remove_empty_rows', 'remove_empty_cols', 'delete_first_X_rows', 'replace_header', 'delete_last_X_rows']
        edit_columns_view_actions = ['add_column', 'delete_columns', 'fill_column', 'split_column', 'merge_columns', 'rename_column']
        edit_data_view_actions = ['delete_data', 'replace_symbol', 'change_case', 'trim_and_replace_multiple_whitespaces']

        # action handlers for summary view
        if action_type in summary_view_actions:
            current_view = 'summary'

            if action_type == 'remove_empty_rows':
                result = handle_remove_empty_rows(df)
                df = result['dataframe']
                
            elif action_type == 'remove_empty_cols':
                cols_to_delete = parameters.get('cols_to_delete')
                result = handle_remove_empty_cols(df, cols_to_delete)
                df = result['dataframe']

                
            elif action_type == 'delete_first_X_rows':
                num_rows_to_delete_start = int(parameters.get('num_rows_to_delete_start', 0))
                result = handle_delete_first_x_rows(df, num_rows_to_delete_start)
                df = result['dataframe']
                
            elif action_type == 'replace_header':
                df = handle_replace_header_with_first_row(df)
                
            elif action_type == 'delete_last_X_rows':
                num_rows_to_delete_end = int(parameters.get('num_rows_to_delete_end', 0))
                result = handle_delete_last_x_rows(df, num_rows_to_delete_end)
                df = result['dataframe']
                

        # action handlers for edit_columns view
        elif action_type in edit_columns_view_actions:
            current_view = 'edit_columns'
        
            if action_type == 'add_column':
                new_column_name = parameters.get('new_column_name')
                df = add_column(df, new_column_name)
                

            elif action_type == 'delete_columns':
                columns_to_delete = parameters.get('columns_to_delete')
                df = delete_columns(df, columns_to_delete)
                

            elif action_type == 'fill_column':
                column_to_fill = parameters.get('column_to_fill')
                fill_value = parameters.get('fill_value')
                fill_option = parameters.get('fill_option')
                if column_to_fill and fill_value is not None:
                    df = fill_column(df, column_to_fill, fill_value, fill_option)
                

            elif action_type == 'split_column':
                column_to_split = parameters.get('column_to_split')
                split_value = parameters.get('split_value')
                delete_original = parameters.get('delete_original')
                ignore_repeated = parameters.get('ignore_repeated', False)
                if column_to_split in df.columns and split_value:
                    df = split_column(df, column_to_split, split_value, delete_original, ignore_repeated)

                

            elif action_type == 'merge_columns':
                columns_to_merge = parameters.get('columns_to_merge')
                merge_separator = parameters.get('merge_separator', '')
                new_column_name = parameters.get('new_column_name')
                delete_original = parameters.get('delete_original', False)

                result = merge_columns(df, columns_to_merge, merge_separator, new_column_name, delete_original)
                df = result['dataframe']

            elif action_type == 'rename_column':
                if is_undo:
                    column_to_rename = parameters.get('new_column_name')
                    new_column_name = parameters.get('column_to_rename')
                else:
                    column_to_rename = parameters.get('column_to_rename')
                    new_column_name = parameters.get('new_column_name')
                if column_to_rename in df.columns and new_column_name:
                    df = rename_column(df, column_to_rename, new_column_name)
                

        # action handlers for edit_data view
        elif action_type in edit_data_view_actions:
            current_view = 'edit_data'
            if action_type == 'delete_data':
                if is_undo:
                    # checks if there is a backup available
                    backup_path = Action.backup_data_path
                    if backup_path:
                        df_backup = pd.read_csv(backup_path)
                        # logic to restore the original data from df_backup
                        for column in df_backup.columns:
                            df[column] = df_backup[column]
                else:
                    columns_to_modify = parameters.get('columns_to_modify')
                    delimiter = parameters.get('delimiter')
                    delete_option = parameters.get('delete_option')
                    include_delimiter = parameters.get('include_delimiter')
                    case_sensitive = parameters.get('case_sensitive')
                    df = delete_data(df, columns_to_modify, delimiter, delete_option, include_delimiter, case_sensitive)
                

            elif action_type == 'replace_symbol':
                columns_to_replace = parameters.get('columns_to_replace')
                if is_undo:
                    # Swap only when undoing
                    old_symbol = parameters.get('new_symbol')
                    new_symbol = parameters.get('old_symbol')
                else:
                    old_symbol = parameters.get('old_symbol')
                    new_symbol = parameters.get('new_symbol')
                case_sensitive = parameters.get('case_sensitive')

                for column in columns_to_replace:
                    df = replace_symbol(df, columns_to_replace, old_symbol, new_symbol, case_sensitive)

            elif action_type == 'trim_and_replace_whitespaces':
                columns_to_modify = parameters.get('columns_to_modify')
                if columns_to_modify == 'All Columns':
                    columns_to_modify = df.columns.tolist()
                df = trim_and_replace_multiple_whitespaces(df, columns_to_modify)

            elif action_type == 'change_case':
                columns_to_change_case = parameters.get('columns_to_change_case')
                case_type = parameters.get('case_type')
                if columns_to_change_case == 'All Columns':
                    columns_to_change_case = df.columns.tolist()
                df = change_case(df, columns_to_change_case, case_type)
                

        else:
            print(f"Unknown action type: {action_type}")
            return df, None
        
        print(f"DataFrame after applying {action_type}:", df.head())
        return df, current_view
    
    except Exception as e:
        print(f"Error applying action {action_type} with parameters {parameters}: {str(e)}")
        return df, str(e)  # returns the df and the error message

@login_required(login_url="/login/")
def save_template(request):
    if request.method == 'POST':
        original_file_path = request.session.get('file_path')
        relative_path = os.path.relpath(original_file_path, settings.MEDIA_ROOT).replace('\\', '/')
        try:
            current_file = UploadedFile.objects.get(file=relative_path)
            df = load_dataframe_from_file(original_file_path)  
            original_headers = df.columns.tolist()  # extracts headers from the df

            actions = get_actions_for_session(request.session.session_key, current_file)
            actions_data = [action_to_dict(action) for action in actions if not action.undone]  # convert sactions to a dict representation

            # checks if actions_data is not empty
            if not actions_data:
                messages.error(request, "No actions to save in the template.")
                return redirect('download')

            template_name = request.POST.get('template_name')
            # creates and saves the template with actions_data and original_headers
            template = Template.objects.create(
                name=template_name, 
                user=request.user, 
                actions=actions_data, 
                original_headers=original_headers
            )
            messages.success(request, f'Template "{template_name}" saved.')
        except UploadedFile.DoesNotExist:
            messages.error(request, "File not found.")
        return redirect('download')
    else:
        messages.error(request, 'Invalid request')
        return redirect('download')
